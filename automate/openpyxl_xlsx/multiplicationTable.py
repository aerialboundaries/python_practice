from sys import argv

import openpyxl
from openpyxl.styles import Font


if len(argv) < 1:
    print('Please give number')
    exit()

n = int(argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
# Write header
for i in range(1, n+1):
    sheet.cell(row=1, column=(i+1)).value = i
    sheet.cell(row=1, column=(i+1)).font = Font(bold=True)
# Write table
for i in range(1, n+1):
    sheet.cell(row=(i+1), column=1).value = i
    sheet.cell(row=(i+1), column=1).font = Font(bold=True)
    for j in range(1, n+1):
        sheet.cell(row=(i+1), column=(j+1)).value = i * j



wb.save('multi.xlsx')
