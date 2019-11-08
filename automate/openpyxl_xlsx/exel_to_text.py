import openpyxl


wb = openpyxl.load_workbook('test_to_excel.xlsx')
ws = wb.active
mr = ws.max_row
mc = ws.max_column
for i in range(1, mc+1):
    with open('test{:0>2}.txt'.format(str(i)), 'w') as tFile:
        for j in range(1, mr+1):
            tFile.write(ws.cell(row=j, column=i).value)
