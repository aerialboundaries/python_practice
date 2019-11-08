from sys import argv

import openpyxl


if len(argv) < 4:
    print('Usage: python blankrowinsert.py 3 2 produce.xlsx')
    exit()

insert_row = int(argv[1])
blank_number = int(argv[2])
orig_file = argv[3]

orig_wb = openpyxl.load_workbook(orig_file)
orig_ws = orig_wb.active
targ_wb = openpyxl.Workbook()
targ_ws = targ_wb.active
rs = 0

for r1 in range(1, orig_ws.max_row+1):
    if (r1 == insert_row):
        rs = blank_number
    for c1 in range(1, orig_ws.max_column):
        targ_ws.cell(row=r1+rs, column=c1).value = \
            orig_ws.cell(row=r1, column=c1).value

targ_wb.save('insertedrow.xlsx')
