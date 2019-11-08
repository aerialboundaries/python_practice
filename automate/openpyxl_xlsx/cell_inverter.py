from sys import argv

import pprint
import openpyxl

wb = openpyxl.load_workbook(argv[1])
ws = wb.active
# read_list = [[cell.value for cell in row] for row in ws.iter_rows()]
# # pprint.pprint(read_list)
# read_transposed_tuple = list(zip(*read_list))
# # pprint.pprint(read_transposed_tuple)
# wb_transposed = openpyxl.Workbook()
# ws_t = wb_transposed.active
# for row in read_transposed_tuple:
#     ws_t.append(row)
# wb_transposed.save(argv[1] + '_transposed.xlsx')

wb_transposed = openpyxl.Workbook()
ws_t = wb_transposed.active

list_t = [[cell.value for cell in cols] for cols in ws.iter_cols()]
for row in list_t:
    ws_t.append(row)    
wb_transposed.save(argv[1] + '_transposed.xlsx')
